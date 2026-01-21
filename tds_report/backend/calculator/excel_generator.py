"""
Excel Report Generator for TDS Calculator
Clean format with one row per transaction
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from io import BytesIO
from typing import List, Dict
import re


# Column headers in exact order
COLUMN_HEADERS = [
    "Deductee Name",
    "Deductee PAN",
    "Detected Category",
    "TDS Section",
    "Transaction Amount",
    "Applicable TDS Rate",
    "TDS Amount",
    "Date of Deduction",
    "Due Date for Payment",
    "Actual Payment Date",
    "Interest Rate",
    "Number of Months",
    "Interest Amount",
    "Total Amount Payable"
]

# Column widths for better readability
COLUMN_WIDTHS = [25, 15, 20, 15, 20, 18, 18, 18, 18, 18, 15, 18, 18, 20]


def generate_excel_report(entity_name: str, pan_number: str, results: List[Dict]) -> BytesIO:
    """
    Generate clean Excel report with one row per transaction.
    
    Args:
        entity_name: Name of the deductor (TAN holder)
        pan_number: TAN number of the deductor
        results: List of calculation result dictionaries
    
    Returns:
        BytesIO object containing the Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "TDS Report"
    
    # Define styles
    header_fill = PatternFill(start_color="1E3C72", end_color="1E3C72", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    data_font = Font(size=10)
    data_alignment = Alignment(horizontal='left', vertical='center')
    currency_alignment = Alignment(horizontal='right', vertical='center')
    
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    # Set column widths
    for col_idx, width in enumerate(COLUMN_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    # Write header row
    ws.row_dimensions[1].height = 30
    for col_idx, header in enumerate(COLUMN_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Write data rows
    for row_idx, result in enumerate(results, 2):
        # Extract data for each column
        row_data = [
            result.get('deductee_name', 'N/A'),
            result.get('deductee_pan', 'N/A'),
            result.get('detected_category', result.get('category_short', '')),
            result.get('section', ''),
            result.get('amount', 0),
            result.get('rate_display', ''),
            result.get('tds_amount', 0),
            result.get('deduction_date_formatted', ''),
            result.get('due_date_formatted', ''),
            result.get('payment_date_formatted', ''),
            "1.5% per month" if result.get('is_late') else "N/A",
            f"{result.get('months_late', 0)} month(s)" if result.get('is_late') else "N/A",
            result.get('interest', 0) if result.get('is_late') else 0,
            result.get('total_payable', 0)
        ]
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.border = thin_border
            
            # Format currency columns (5, 7, 13, 14)
            if col_idx in [5, 7, 13, 14]:
                cell.alignment = currency_alignment
                if isinstance(value, (int, float)):
                    cell.number_format = '₹#,##0.00'
            else:
                cell.alignment = data_alignment
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def get_excel_filename(entity_name: str = None) -> str:
    """
    Generate Excel filename using entity name.
    """
    if not entity_name:
        entity_name = 'TDS_Report'
    
    # Sanitize name for filename
    safe_name = re.sub(r'[^\w\s-]', '', entity_name).strip().replace(' ', '_')
    date_str = datetime.now().strftime('%Y%m%d')
    return f"{safe_name}_TDS_Report_{date_str}.xlsx"
